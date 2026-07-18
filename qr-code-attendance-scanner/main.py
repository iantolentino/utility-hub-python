import tkinter as tk
from tkinter import ttk
from datetime import datetime
import threading
import winsound  # For beep sound (Windows only)
import os
from openpyxl import Workbook, load_workbook

class QRCodeScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Code Scanner")

        # Center the window and set size to 800x600.
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = 800
        window_height = 600
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        root.configure(bg='white')

        # Each session gets a unique ID.
        self.session_id = datetime.now().strftime("%Y%m%d%H%M%S")
        # Dictionary to store scanned data.
        # Structure: self.scanned_data[name] is a list of dicts {date, id, time_in, time_out}
        self.scanned_data = {}
        self.scan_lock = False

        # UI Elements
        self.details_label = tk.Label(
            root,
            text="Align QR code and wait for the beep.",
            font=("Arial", 24),
            fg="gray",
            bg="white"
        )
        self.details_label.pack(pady=10)

        # Date Label.
        self.date_label = tk.Label(
            root,
            text=self.get_formatted_date(),
            font=("Arial", 24),
            fg="black",
            bg="white"
        )
        self.date_label.pack(pady=10)

        self.input_entry = tk.Entry(root)
        self.input_entry.config(
            width=1,
            borderwidth=0,
            highlightthickness=0,
            bg="white",
            fg="white",
            insertbackground="white"
        )
        self.input_entry.pack(pady=0)
        self.input_entry.bind("<Return>", self.on_scan_complete)
        self.input_entry.focus_set()

        # TreeView for displaying scanned data.
        self.tree = ttk.Treeview(
            root,
            columns=("Name", "ID", "Time In", "Time Out"),
            show='headings',
            displaycolumns=("Name", "ID", "Time In", "Time Out"),
            height=15
        )
        for col in ("Name", "ID", "Time In", "Time Out"):
            self.tree.heading(col, text=col, anchor="center")
            if col == "ID":
                self.tree.column(col, anchor="center", width=150, stretch=True)
            else:
                self.tree.column(col, anchor="center", width=200, stretch=True)
        self.tree.pack(pady=20, expand=True, fill='both')

        # Style for the TreeView.
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background='white', fieldbackground='white', foreground='black')
        style.configure('Treeview.Heading', background='lightgray', foreground='black', font=('Arial', 18))

        # Notification label.
        self.notification_label = tk.Label(root, text="", font=("Arial", 24), fg="green", bg="white")
        self.notification_label.pack(pady=10)

        self.update_date_label()

    def get_formatted_date(self):
        now = datetime.now()
        return now.strftime("%B %d, %Y (%I:%M %p)")

    def update_date_label(self):
        self.date_label.config(text=self.get_formatted_date())
        self.root.after(60000, self.update_date_label)

    def on_scan_complete(self, event):
        """
        Processes a scan when Enter is pressed.
        Expected format: "Name: <Name>, ID: <IDNumber>".
        """
        text = self.input_entry.get().strip()
        parts = text.split(",")
        if len(parts) == 2 and not self.scan_lock:
            name_part = parts[0].strip()
            id_part = parts[1].strip()
            if name_part.lower().startswith("name:") and id_part.lower().startswith("id:"):
                name = name_part[5:].strip()
                uid = id_part[3:].strip()
                today_str = datetime.now().strftime("%Y-%m-%d")

                if name not in self.scanned_data:
                    self.scanned_data[name] = []

                # Get today's entries for this name.
                today_entries = [e for e in self.scanned_data[name] if e["date"] == today_str]
                entry = today_entries[-1] if today_entries else None

                self.scan_lock = True
                threading.Timer(1.5, self.release_scan_lock).start()

                if entry is None:
                    # Record Time In.
                    new_entry = {
                        "date": today_str,
                        "id": uid,
                        "time_in": datetime.now().strftime("%H:%M:%S"),
                        "time_out": None
                    }
                    self.scanned_data[name].append(new_entry)
                    winsound.Beep(1000, 300)
                    self.show_notification_popup(f"{name}: Time In recorded.", "green")
                elif entry["time_out"] is None:
                    # Record Time Out.
                    entry["time_out"] = datetime.now().strftime("%H:%M:%S")
                    winsound.Beep(1500, 300)
                    self.show_notification_popup(f"{name}: Time Out recorded.", "green")
                else:
                    winsound.Beep(800, 200)
                    self.show_notification_popup(f"{name}: Already scanned twice today.", "red")

                self.update_table()
                self.save_to_excel()

        # Clear input field.
        self.root.after(100, lambda: self.input_entry.delete(0, tk.END))

    def release_scan_lock(self):
        self.scan_lock = False

    def show_notification_popup(self, msg, color="green"):
        self.notification_label.config(text=msg, fg=color)
        self.root.after(2000, lambda: self.notification_label.config(text=""))

    def update_table(self):
        """Refreshes the TreeView with today’s data only."""
        today_str = datetime.now().strftime("%Y-%m-%d")
        for row in self.tree.get_children():
            self.tree.delete(row)
        for name, entries in self.scanned_data.items():
            for e in entries:
                if e["date"] == today_str:
                    self.tree.insert("", "end", values=(
                        name,
                        e["id"],
                        e["time_in"],
                        e["time_out"] if e["time_out"] else ""
                    ))

    def save_to_excel(self):
        """
        Writes (or updates) the current session block at the top of the Excel file.
        Each session block has a unique marker based on self.session_id.
        If a session block from the current session is already present at the top, it is removed
        so that only one block for the current session appears. Blocks from earlier sessions remain.
        """
        folder_path = r"C:\AttendanceSystem\daily_data"
        file_path = os.path.join(folder_path, "attendance_copy.xlsx")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Attendance Records"

        # Remove existing current session block (if present).
        self._remove_top_block_current_session(sheet)

        # Build the new block.
        # The header will include our session marker and a visible timestamp.
        timestamp = datetime.now().strftime("%m/%d/%Y - %I:%M %p")
        header_text = f"##SESSION_{self.session_id}##"
        today_str = datetime.now().strftime("%Y-%m-%d")
        block_rows = []
        block_rows.append([header_text])
        block_rows.append([timestamp])
        block_rows.append(["Name", "ID", "Time In", "Time Out"])

        # Gather today's entries (sort them by time_in for order).
        today_entries = []
        for name, entries in self.scanned_data.items():
            for e in entries:
                if e["date"] == today_str:
                    today_entries.append((e["time_in"], name, e["id"], e["time_in"], e["time_out"] if e["time_out"] else ""))
        today_entries.sort(key=lambda x: x[0])
        for _, name, id_, in_time, out_time in today_entries:
            block_rows.append([name, id_, in_time, out_time])

        # Add a blank separator row.
        block_rows.append([])

        # Insert the block at the top.
        num_new_rows = len(block_rows)
        sheet.insert_rows(1, num_new_rows)
        for i, row_data in enumerate(block_rows, start=1):
            for j, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=i, column=j, value=cell_value)

        workbook.save(file_path)

    def _remove_top_block_current_session(self, sheet):
        """
        Checks whether the very top block (starting at row 1) was written for the current session
        by looking for the marker "##SESSION_<session_id>##". If found, remove that entire block.
        The block is assumed to end at the first blank row.
        """
        if sheet.max_row < 1:
            return

        top_cell_value = sheet.cell(row=1, column=1).value
        marker = f"##SESSION_{self.session_id}##"
        if top_cell_value and isinstance(top_cell_value, str) and top_cell_value.startswith(marker):
            # Found the current session block at the top.
            row_idx = 1
            while row_idx <= sheet.max_row:
                cell_val = sheet.cell(row=row_idx, column=1).value
                # Remove until a blank row is encountered (or end of block).
                if not cell_val:
                    sheet.delete_rows(row_idx, 1)
                    break
                else:
                    sheet.delete_rows(row_idx, 1)
            # (No need to adjust row_idx since rows shift up on deletion.)

if __name__ == "__main__":
    root = tk.Tk()
    app = QRCodeScannerApp(root)
    root.mainloop()
