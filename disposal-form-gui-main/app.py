"""
Disposal Form GUI (single-file) — Updated
- Centers the window on screen.
- When "Save Session" is clicked, the session is immediately appended to the Excel file.
- Shows the full path to the Excel file in the UI so you can open/check it.
- Minimal, modern-ish tkinter + ttk UI (no popup confirmation on save; status shown in-app).
Dependencies:
    pip install openpyxl
Run:
    python disposal_form.py
"""

import tkinter as tk
from tkinter import ttk
from datetime import datetime
import os
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except Exception:
    raise SystemExit("Missing dependency: openpyxl. Install with: pip install openpyxl")

# ---------- Config ----------
CATEGORIES = ["White paper", "Colored Paper", "Plastic", "Garbage"]
UNIT_OPTIONS = ["KG", "Grams"]

# default Excel path: Documents/disposal_log.xlsx if Documents exists, otherwise home
home = Path.home()
documents = home / "Documents"
if documents.exists():
    EXCEL_FILENAME = str(documents / "disposal_log.xlsx")
else:
    EXCEL_FILENAME = str(home / "disposal_log.xlsx")


# ---------- Excel utilities ----------
def ensure_workbook(path):
    """Return openpyxl Workbook object (loaded or new)"""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Disposal Log"
        wb.save(path)
    return wb


def append_session_to_excel(path, timestamp_str, rows, total_kg):
    """
    rows: list of tuples (category, quantity_as_entered, unit_as_entered)
    total_kg: float
    Writes a block:
    Date/Time added: <timestamp>
    Category | Quantity | Unit
    ...
    Total: <total_kg> KG
    ...
    """
    wb = ensure_workbook(path)
    ws = wb.active

    # find next empty row (ws.max_row is safe)
    next_row = ws.max_row + 1

    # timestamp row (bold)
    ts_cell = ws.cell(row=next_row, column=1, value=f"Date/Time Added: {timestamp_str}")
    ts_cell.font = Font(bold=True)
    ts_cell.alignment = Alignment(horizontal="left")
    next_row += 1

    # header
    h_cat = ws.cell(row=next_row, column=1, value="Category")
    h_qty = ws.cell(row=next_row, column=2, value="Quantity")
    h_unit = ws.cell(row=next_row, column=3, value="Unit")
    for c in (h_cat, h_qty, h_unit):
        c.font = Font(bold=True)
    next_row += 1

    # rows
    for cat, qty_text, unit in rows:
        ws.cell(row=next_row, column=1, value=cat)
        try:
            qty_num = float(qty_text)
            ws.cell(row=next_row, column=2, value=qty_num)
        except Exception:
            ws.cell(row=next_row, column=2, value=qty_text)
        ws.cell(row=next_row, column=3, value=unit)
        next_row += 1

    # total row
    total_cell_label = ws.cell(row=next_row, column=1, value="Total: ")
    total_cell_value = ws.cell(row=next_row, column=2, value=round(total_kg, 3))
    total_cell_unit = ws.cell(row=next_row, column=3, value="KG")
    for c in (total_cell_label, total_cell_value, total_cell_unit):
        c.font = Font(bold=True)

    # add one blank row below the total (merged across 3 columns) and make it slightly taller
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=3)
    # leave the merged cell empty; set height to create visible spacing (tweak value as needed)
    ws.row_dimensions[next_row].height = 12

    # advance past the blank row so subsequent content doesn't overwrite it
    next_row += 2

    wb.save(path)


# ---------- Helpers ----------
def parse_quantity_to_kg(q_str, unit):
    """Parse quantity string and convert to KG (float). Returns 0.0 on failure."""
    try:
        q = float(q_str)
    except Exception:
        return 0.0
    if unit == "KG":
        return q
    elif unit == "Grams":
        return q / 1000.0
    else:
        return q


# ---------- GUI ----------
class DisposalApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Disposal Form")
        self.width = 520
        self.height = 420
        self.minsize(480, 380)

        # center & set geometry
        self.geometry(f"{self.width}x{self.height}")
        self.center_window(self.width, self.height)

        # ttk style
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TLabel", font=("Segoe UI", 10))
        style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"))
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("TEntry", font=("Segoe UI", 10))
        style.configure("TCombobox", font=("Segoe UI", 10))

        self.session_time = tk.StringVar()
        self.total_kg_var = tk.StringVar(value="0.000 KG")
        self.status_var = tk.StringVar(value="")
        self.file_path_var = tk.StringVar(value=EXCEL_FILENAME)

        self.qty_vars = {}
        self.unit_vars = {}

        self._build_ui()
        self.new_session()

    def center_window(self, width, height):
        self.update_idletasks()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = (screen_w - width) // 2
        y = (screen_h - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")

    def _build_ui(self):
        padding = {"padx": 12, "pady": 6}

        header = ttk.Label(self, text="Disposal Form", style="Header.TLabel")
        header.pack(anchor="w", padx=16, pady=(12, 0))

        sess_frame = ttk.Frame(self)
        sess_frame.pack(fill="x", padx=12, pady=(6, 8))

        ttk.Label(sess_frame, text="Session timestamp:").grid(row=0, column=0, sticky="w")
        ttk.Label(sess_frame, textvariable=self.session_time).grid(row=0, column=1, sticky="w", padx=(8, 0))

        ttk.Label(sess_frame, text="Live total:").grid(row=0, column=2, sticky="e", padx=(20, 0))
        ttk.Label(sess_frame, textvariable=self.total_kg_var).grid(row=0, column=3, sticky="w", padx=(8, 0))

        # Categories form
        form_frame = ttk.Frame(self)
        form_frame.pack(fill="both", expand=False, padx=12, pady=(4, 8))

        ttk.Label(form_frame, text="Category").grid(row=0, column=0, sticky="w", padx=(2, 10))
        ttk.Label(form_frame, text="Quantity").grid(row=0, column=1, sticky="w", padx=(2, 10))
        ttk.Label(form_frame, text="Unit").grid(row=0, column=2, sticky="w", padx=(2, 10))

        for i, cat in enumerate(CATEGORIES, start=1):
            ttk.Label(form_frame, text=cat).grid(row=i, column=0, sticky="w", padx=(2, 10), pady=6)

            v = tk.StringVar(value="0")
            e = ttk.Entry(form_frame, textvariable=v, width=12)
            e.grid(row=i, column=1, sticky="w")
            v.trace_add("write", self._on_qty_change)
            self.qty_vars[cat] = v

            uv = tk.StringVar(value="KG")
            cb = ttk.Combobox(form_frame, values=UNIT_OPTIONS, state="readonly", width=8, textvariable=uv)
            cb.grid(row=i, column=2, sticky="w", padx=(6, 0))
            cb.bind("<<ComboboxSelected>>", lambda e: self._on_qty_change())
            self.unit_vars[cat] = uv

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=12, pady=(8, 6))

        new_btn = ttk.Button(btn_frame, text="New Session", command=self.new_session)
        new_btn.pack(side="left")

        save_btn = ttk.Button(btn_frame, text="Save Session (Add new data)", command=self.save_session)
        save_btn.pack(side="left", padx=8)

        open_btn = ttk.Button(btn_frame, text="Open Excel File", command=self.open_excel)
        open_btn.pack(side="right")

        # file path & status area
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill="x", padx=12, pady=(6, 12))

        ttk.Label(bottom_frame, text="Excel file:").grid(row=0, column=0, sticky="w")
        ttk.Label(bottom_frame, textvariable=self.file_path_var).grid(row=0, column=1, sticky="w", padx=(6, 0))

        ttk.Label(bottom_frame, textvariable=self.status_var).grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))

        footer = ttk.Label(self, text="Units: KG or Grams. Quantities will be summed and stored in KG.", font=("Segoe UI", 9))
        footer.pack(side="bottom", fill="x", padx=12, pady=(0, 8))

    def new_session(self):
        now = datetime.now()
        self.session_timestamp = now
        self.session_time.set(now.strftime("%Y-%m-%d %H:%M:%S"))
        for cat in CATEGORIES:
            self.qty_vars[cat].set("0")
            self.unit_vars[cat].set("KG")
        self._update_total_label()
        self.status_var.set("")

    def _on_qty_change(self, *args):
        self._update_total_label()

    def _update_total_label(self):
        total = 0.0
        for cat in CATEGORIES:
            qstr = self.qty_vars[cat].get().strip()
            unit = self.unit_vars[cat].get()
            kg = parse_quantity_to_kg(qstr, unit)
            total += kg
        self.total_kg_var.set(f"{round(total, 3):.3f} KG")

    def save_session(self):
        # Gather input rows
        rows = []
        for cat in CATEGORIES:
            qstr = self.qty_vars[cat].get().strip()
            unit = self.unit_vars[cat].get()
            if qstr == "":
                qstr = "0"
            rows.append((cat, qstr, unit))

        total = sum(parse_quantity_to_kg(q, u) for _, q, u in rows)
        timestamp_str = self.session_time.get()

        try:
            append_session_to_excel(EXCEL_FILENAME, timestamp_str, rows, total)
        except Exception as e:
            # show immediate status in-app (no popup)
            self.status_var.set(f"Save failed: {e}")
            return

        # update status and the file path display so user can verify
        self.file_path_var.set(EXCEL_FILENAME)
        self.status_var.set(f"Saved to: {EXCEL_FILENAME} — Total: {round(total,3):.3f} KG")
        # automatically start a new session (as requested by workflow)
        self.new_session()

    def open_excel(self):
        if not os.path.exists(EXCEL_FILENAME):
            self.status_var.set("Excel file not found yet. Save a session first.")
            return
        try:
            if os.name == "nt":
                os.startfile(EXCEL_FILENAME)
            elif os.name == "posix":
                import subprocess
                # macOS uses "open", linux commonly "xdg-open"
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.Popen([opener, EXCEL_FILENAME])
            else:
                self.status_var.set(f"File saved at: {os.path.abspath(EXCEL_FILENAME)}")
        except Exception:
            self.status_var.set(f"File saved at: {os.path.abspath(EXCEL_FILENAME)}")


if __name__ == "__main__":
    app = DisposalApp()
    app.mainloop()
